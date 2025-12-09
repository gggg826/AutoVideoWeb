"""
管理后台 API 路由
提供访问数据查询、统计分析等功能
"""
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, desc
from app.core.database import get_db
from app.models.visit import Visit
from app.schemas.visit import VisitDetail
from app.api.v1.auth import get_current_admin
from typing import List, Optional
from datetime import datetime, timedelta
import csv
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

router = APIRouter(
    prefix="/admin",
    tags=["admin"],
    dependencies=[Depends(get_current_admin)]
)


@router.get("/visits", summary="获取访问列表")
async def get_visits(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=10000, description="每页��量"),
    device_type: Optional[str] = Query(None, description="设备类型筛选"),
    min_score: Optional[float] = Query(None, ge=0, le=100, description="最低评分"),
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db)
):
    """
    获取访问记录列表（分页）

    支持筛选条件：
    - device_type: 设备类型（pc/mobile/tablet/bot）
    - min_score: 最低真实性评分
    - start_date/end_date: 时间范围
    """
    # 构建查询条件
    conditions = []

    if device_type:
        conditions.append(Visit.device_type == device_type)

    if min_score is not None:
        conditions.append(Visit.authenticity_score >= min_score)

    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date)
            conditions.append(Visit.timestamp >= start_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="无效的开始日期格式")

    if end_date:
        try:
            end_dt = datetime.fromisoformat(end_date) + timedelta(days=1)
            conditions.append(Visit.timestamp < end_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="无效的结束日期格式")

    # 查询总数
    count_stmt = select(func.count(Visit.id))
    if conditions:
        count_stmt = count_stmt.where(and_(*conditions))

    result = await db.execute(count_stmt)
    total = result.scalar_one()

    # 查询数据
    offset = (page - 1) * page_size
    stmt = select(Visit).order_by(desc(Visit.timestamp)).limit(page_size).offset(offset)

    if conditions:
        stmt = stmt.where(and_(*conditions))

    result = await db.execute(stmt)
    visits = result.scalars().all()

    return {
        "success": True,
        "data": [
            {
                "id": v.id,
                "visit_id": v.visit_id,
                "timestamp": v.timestamp.isoformat() if v.timestamp else None,
                "ip_address": v.ip_address,
                "ip_country": v.ip_country,
                "ip_city": v.ip_city,
                "device_type": v.device_type,
                "browser": v.browser,
                "os": v.os,
                "stay_duration": v.stay_duration,
                "scroll_depth": v.scroll_depth,
                "is_bot": v.is_bot,
                "authenticity_score": v.authenticity_score,
            }
            for v in visits
        ],
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": (total + page_size - 1) // page_size
        }
    }


@router.get("/visits/{visit_id}", summary="获取访问详情")
async def get_visit_detail(
    visit_id: str,
    db: AsyncSession = Depends(get_db)
):
    """获取单个访问记录的详细信息"""
    stmt = select(Visit).where(Visit.visit_id == visit_id)
    result = await db.execute(stmt)
    visit = result.scalar_one_or_none()

    if not visit:
        raise HTTPException(status_code=404, detail="访问记录不存在")

    return {
        "success": True,
        "data": {
            "id": visit.id,
            "visit_id": visit.visit_id,
            "timestamp": visit.timestamp.isoformat() if visit.timestamp else None,

            # IP 信息
            "ip_address": visit.ip_address,
            "ip_country": visit.ip_country,
            "ip_city": visit.ip_city,
            "is_proxy": visit.is_proxy,

            # 请求信息
            "user_agent": visit.user_agent,
            "referrer": visit.referrer,
            "page_url": visit.page_url,

            # 设备信息
            "device_type": visit.device_type,
            "browser": visit.browser,
            "browser_version": visit.browser_version,
            "os": visit.os,
            "os_version": visit.os_version,

            # 浏览器指纹
            "screen_resolution": visit.screen_resolution,
            "timezone": visit.timezone,
            "language": visit.language,
            "platform": visit.platform,
            "canvas_fingerprint": visit.canvas_fingerprint,
            "webgl_fingerprint": visit.webgl_fingerprint,
            "fonts_hash": visit.fonts_hash,

            # 行为数据
            "stay_duration": visit.stay_duration,
            "scroll_depth": visit.scroll_depth,
            "mouse_movements": visit.mouse_movements,

            # 分析字段
            "is_bot": visit.is_bot,
            "authenticity_score": visit.authenticity_score,
            "fingerprint_hash": visit.fingerprint_hash,
        }
    }


@router.get("/stats/summary", summary="获取统计摘要")
async def get_stats_summary(
    days: int = Query(7, ge=1, le=90, description="统计天数"),
    db: AsyncSession = Depends(get_db)
):
    """
    获取统计摘要

    包含：
    - 总访问量
    - 今日访问量
    - 昨日访问量
    - 平均真实性评分
    - 设备类型分布
    - 机器人访问占比
    """
    now = datetime.now()
    today_start = datetime(now.year, now.month, now.day)
    yesterday_start = today_start - timedelta(days=1)
    period_start = today_start - timedelta(days=days)

    # 总访问量
    total_stmt = select(func.count(Visit.id))
    total_result = await db.execute(total_stmt)
    total_visits = total_result.scalar_one()

    # 今日访问量
    today_stmt = select(func.count(Visit.id)).where(Visit.timestamp >= today_start)
    today_result = await db.execute(today_stmt)
    today_visits = today_result.scalar_one()

    # 昨日访问量
    yesterday_stmt = select(func.count(Visit.id)).where(
        and_(
            Visit.timestamp >= yesterday_start,
            Visit.timestamp < today_start
        )
    )
    yesterday_result = await db.execute(yesterday_stmt)
    yesterday_visits = yesterday_result.scalar_one()

    # 周期内访问量
    period_stmt = select(func.count(Visit.id)).where(Visit.timestamp >= period_start)
    period_result = await db.execute(period_stmt)
    period_visits = period_result.scalar_one()

    # 平均真实性评分
    avg_score_stmt = select(func.avg(Visit.authenticity_score)).where(Visit.timestamp >= period_start)
    avg_score_result = await db.execute(avg_score_stmt)
    avg_score = avg_score_result.scalar_one() or 0.0

    # 设备类型分布
    device_stmt = select(
        Visit.device_type,
        func.count(Visit.id).label('count')
    ).where(Visit.timestamp >= period_start).group_by(Visit.device_type)
    device_result = await db.execute(device_stmt)
    device_distribution = {row[0] or 'unknown': row[1] for row in device_result}

    # 机器人访问数
    bot_stmt = select(func.count(Visit.id)).where(
        and_(
            Visit.timestamp >= period_start,
            Visit.is_bot == True
        )
    )
    bot_result = await db.execute(bot_stmt)
    bot_visits = bot_result.scalar_one()

    return {
        "success": True,
        "data": {
            "total_visits": total_visits,
            "today_visits": today_visits,
            "yesterday_visits": yesterday_visits,
            "period_visits": period_visits,
            "period_days": days,
            "avg_authenticity_score": round(avg_score, 2),
            "device_distribution": device_distribution,
            "bot_visits": bot_visits,
            "bot_rate": round(bot_visits / period_visits * 100, 2) if period_visits > 0 else 0,
        }
    }


@router.get("/stats/trend", summary="获取访问趋势")
async def get_stats_trend(
    days: int = Query(7, ge=1, le=90, description="统计天数"),
    db: AsyncSession = Depends(get_db)
):
    """
    获取访问趋势数据（按天统计）
    """
    period_start = datetime.now() - timedelta(days=days)

    # 按天统计访问量（SQLite 兼容）
    stmt = select(
        func.date(Visit.timestamp).label('date'),
        func.count(Visit.id).label('visits'),
        func.avg(Visit.authenticity_score).label('avg_score')
    ).where(
        Visit.timestamp >= period_start
    ).group_by(
        func.date(Visit.timestamp)
    ).order_by('date')

    result = await db.execute(stmt)
    rows = result.all()

    return {
        "success": True,
        "data": [
            {
                "date": str(row[0]) if row[0] else None,
                "visits": row[1],
                "avg_score": round(row[2], 2) if row[2] else 0
            }
            for row in rows
        ]
    }


@router.get("/stats/devices", summary="获取设备统计")
async def get_device_stats(
    days: int = Query(7, ge=1, le=90, description="统计天数"),
    db: AsyncSession = Depends(get_db)
):
    """获取设备类型、浏览器、操作系统统计"""
    period_start = datetime.now() - timedelta(days=days)

    # 设备类型统计
    device_stmt = select(
        Visit.device_type,
        func.count(Visit.id).label('count')
    ).where(Visit.timestamp >= period_start).group_by(Visit.device_type)
    device_result = await db.execute(device_stmt)

    # 浏览器统计
    browser_stmt = select(
        Visit.browser,
        func.count(Visit.id).label('count')
    ).where(
        and_(
            Visit.timestamp >= period_start,
            Visit.browser != None
        )
    ).group_by(Visit.browser).order_by(desc('count')).limit(10)
    browser_result = await db.execute(browser_stmt)

    # 操作系统统计
    os_stmt = select(
        Visit.os,
        func.count(Visit.id).label('count')
    ).where(
        and_(
            Visit.timestamp >= period_start,
            Visit.os != None
        )
    ).group_by(Visit.os).order_by(desc('count')).limit(10)
    os_result = await db.execute(os_stmt)

    return {
        "success": True,
        "data": {
            "devices": [
                {"name": row[0] or 'unknown', "count": row[1]}
                for row in device_result
            ],
            "browsers": [
                {"name": row[0], "count": row[1]}
                for row in browser_result
            ],
            "operating_systems": [
                {"name": row[0], "count": row[1]}
                for row in os_result
            ]
        }
    }


@router.get("/stats/locations", summary="获取地理位置统计")
async def get_location_stats(
    days: int = Query(7, ge=1, le=90, description="统计天数"),
    db: AsyncSession = Depends(get_db)
):
    """获取国家和城市访问统计"""
    period_start = datetime.now() - timedelta(days=days)

    # 国家统计
    country_stmt = select(
        Visit.ip_country,
        func.count(Visit.id).label('count')
    ).where(
        and_(
            Visit.timestamp >= period_start,
            Visit.ip_country != None
        )
    ).group_by(Visit.ip_country).order_by(desc('count')).limit(15)
    country_result = await db.execute(country_stmt)

    # 城市统计（前15个）
    city_stmt = select(
        Visit.ip_country,
        Visit.ip_city,
        func.count(Visit.id).label('count')
    ).where(
        and_(
            Visit.timestamp >= period_start,
            Visit.ip_city != None
        )
    ).group_by(Visit.ip_country, Visit.ip_city).order_by(desc('count')).limit(15)
    city_result = await db.execute(city_stmt)

    return {
        "success": True,
        "data": {
            "countries": [
                {"code": row[0], "count": row[1]}
                for row in country_result
            ],
            "cities": [
                {"country": row[0], "city": row[1], "count": row[2]}
                for row in city_result
            ]
        }
    }


@router.get("/stats/referrers", summary="获取来源渠道统计")
async def get_referrer_stats(
    days: int = Query(7, ge=1, le=90, description="统计天数"),
    db: AsyncSession = Depends(get_db)
):
    """获取来源渠道/平台统计"""
    from urllib.parse import urlparse

    period_start = datetime.now() - timedelta(days=days)

    # 来源统计 - 对referrer进行分类处理
    referrer_stmt = select(
        Visit.referrer,
        func.count(Visit.id).label('count')
    ).where(
        Visit.timestamp >= period_start
    ).group_by(Visit.referrer).order_by(desc('count')).limit(20)

    referrer_result = await db.execute(referrer_stmt)
    referrers_raw = [(row[0], row[1]) for row in referrer_result]

    # 分类处理来源
    categorized = {}
    direct_count = 0

    for referrer, count in referrers_raw:
        if not referrer or referrer == "" or referrer == "null" or referrer == "None":
            direct_count += count
        else:
            # 简单的域名提取和分类
            try:
                parsed = urlparse(referrer)
                domain = parsed.netloc.lower() if parsed.netloc else ""

                # 如果没有域名，归入直接访问
                if not domain:
                    direct_count += count
                    continue

                # 去除www前缀
                if domain.startswith('www.'):
                    domain = domain[4:]

                # 分类主要平台
                if 'google' in domain:
                    category = 'Google'
                elif 'bing' in domain or 'microsoft' in domain:
                    category = 'Bing'
                elif 'baidu' in domain:
                    category = '百度'
                elif 'tiktok' in domain or 'douyin' in domain:
                    category = 'TikTok/抖音'
                elif 'facebook' in domain or 'fb.com' in domain:
                    category = 'Facebook'
                elif 'twitter' in domain or 'x.com' in domain:
                    category = 'Twitter/X'
                elif 'youtube' in domain:
                    category = 'YouTube'
                elif 'instagram' in domain:
                    category = 'Instagram'
                elif 'weibo' in domain:
                    category = '微博'
                elif 'wechat' in domain or 'weixin' in domain:
                    category = '微信'
                elif 'linkedin' in domain:
                    category = 'LinkedIn'
                elif 'reddit' in domain:
                    category = 'Reddit'
                else:
                    # 其他来源使用域名
                    category = domain if len(domain) < 30 else domain[:27] + '...'

                if category in categorized:
                    categorized[category] += count
                else:
                    categorized[category] = count
            except Exception as e:
                # 解析失败的归入其他
                if '其他' in categorized:
                    categorized['其他'] += count
                else:
                    categorized['其他'] = count

    # 添加直接访问
    if direct_count > 0:
        categorized['直接访问'] = direct_count

    # 如果没有任何数据，返回空列表
    if not categorized:
        return {
            "success": True,
            "data": {
                "referrers": []
            }
        }

    # 转换为列表并排序
    referrers_list = [
        {"name": name, "count": count}
        for name, count in sorted(categorized.items(), key=lambda x: x[1], reverse=True)
    ]

    return {
        "success": True,
        "data": {
            "referrers": referrers_list
        }
    }


@router.get("/stats/hourly-admin", summary="获取管理员本地时间访问分布")
async def get_hourly_admin_stats(db: AsyncSession = Depends(get_db)):
    """
    获取最近24小时的访问分布（管理员本地时间）
    返回原始时间戳，由前端转换为浏览器本地时间
    """
    # 获取最近24小时的访问记录
    time_24h_ago = datetime.utcnow() - timedelta(hours=24)

    stmt = select(Visit.timestamp).where(
        Visit.timestamp >= time_24h_ago
    ).order_by(Visit.timestamp)

    result = await db.execute(stmt)
    timestamps = [row[0] for row in result]

    # 将datetime对象转换为ISO格式字符串
    timestamps_iso = [ts.isoformat() if ts else None for ts in timestamps]

    return {
        "success": True,
        "data": {
            "timestamps": timestamps_iso
        }
    }


@router.get("/stats/hourly-visitor", summary="获取访问者本地时间访问分布")
async def get_hourly_visitor_stats(
    country: Optional[str] = Query(None, description="国家代码筛选"),
    db: AsyncSession = Depends(get_db)
):
    """
    获取最近24小时的访问分布（访问者本地时间）
    根据访问者的timezone和country推算其本地时间
    """
    try:
        from zoneinfo import ZoneInfo
    except ImportError:
        # Python 3.8 fallback
        from backports.zoneinfo import ZoneInfo

    # 国家到时区的映射（主要时区）
    country_timezone_map = {
        'US': 'America/New_York',
        'CN': 'Asia/Shanghai',
        'JP': 'Asia/Tokyo',
        'KR': 'Asia/Seoul',
        'GB': 'Europe/London',
        'FR': 'Europe/Paris',
        'DE': 'Europe/Berlin',
        'AU': 'Australia/Sydney',
        'CA': 'America/Toronto',
        'IN': 'Asia/Kolkata',
        'BR': 'America/Sao_Paulo',
        'RU': 'Europe/Moscow',
        'SG': 'Asia/Singapore',
        'HK': 'Asia/Hong_Kong',
        'TW': 'Asia/Taipei',
    }

    # 获取最近24小时的访问记录
    time_24h_ago = datetime.utcnow() - timedelta(hours=24)

    stmt = select(Visit.timestamp, Visit.timezone, Visit.ip_country).where(
        Visit.timestamp >= time_24h_ago
    )

    # 如果指定了国家，则筛选
    if country:
        stmt = stmt.where(Visit.ip_country == country)

    stmt = stmt.order_by(Visit.timestamp)

    result = await db.execute(stmt)
    records = [(row[0], row[1], row[2]) for row in result]

    # 转换为访问者本地时间
    local_timestamps = []
    for timestamp, visitor_tz, ip_country in records:
        if not timestamp:
            continue

        try:
            # 优先使用访问者的timezone字段
            if visitor_tz:
                # 尝试解析timezone字符串（如 "Asia/Shanghai"）
                try:
                    tz = ZoneInfo(visitor_tz)
                    # 假设timestamp是aware datetime (UTC)
                    if timestamp.tzinfo is None:
                        timestamp = timestamp.replace(tzinfo=ZoneInfo('UTC'))
                    local_time = timestamp.astimezone(tz)
                    local_timestamps.append(local_time.isoformat())
                    continue
                except:
                    pass

            # 如果没有timezone字段，使用国家映射
            if ip_country and ip_country in country_timezone_map:
                tz = ZoneInfo(country_timezone_map[ip_country])
                if timestamp.tzinfo is None:
                    timestamp = timestamp.replace(tzinfo=ZoneInfo('UTC'))
                local_time = timestamp.astimezone(tz)
                local_timestamps.append(local_time.isoformat())
            else:
                # 无法确定时区，使用UTC
                local_timestamps.append(timestamp.isoformat())
        except Exception as e:
            # 出错时使用原始时间
            local_timestamps.append(timestamp.isoformat() if timestamp else None)

    return {
        "success": True,
        "data": {
            "timestamps": local_timestamps,
            "country": country
        }
    }


@router.post("/clear-visits", summary="清空所有访问记录")
async def clear_all_visits(db: AsyncSession = Depends(get_db)):
    """
    清空所有访问记录

    警告：此操作不可逆！
    """
    try:
        # 获取删除前的记录数
        count_stmt = select(func.count(Visit.id))
        result = await db.execute(count_stmt)
        total_count = result.scalar()

        # 删除所有记录
        delete_stmt = Visit.__table__.delete()
        await db.execute(delete_stmt)
        await db.commit()

        return {
            "success": True,
            "message": f"已清空所有访问记录",
            "deleted_count": total_count
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"清空失败: {str(e)}")


@router.delete("/visits/{visit_id}", summary="删除访问记录")
async def delete_visit(
    visit_id: str,
    db: AsyncSession = Depends(get_db)
):
    """删除单个访问记录"""
    stmt = select(Visit).where(Visit.visit_id == visit_id)
    result = await db.execute(stmt)
    visit = result.scalar_one_or_none()

    if not visit:
        raise HTTPException(status_code=404, detail="访问记录不存在")

    await db.delete(visit)
    await db.commit()

    return {
        "success": True,
        "message": "访问记录已删除"
    }


@router.get("/export/csv", summary="导出 CSV")
async def export_csv(
    device_type: Optional[str] = Query(None, description="设备类型筛选"),
    min_score: Optional[float] = Query(None, ge=0, le=100, description="最低评分"),
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    limit: int = Query(10000, ge=1, le=50000, description="最大导出数量"),
    db: AsyncSession = Depends(get_db)
):
    """
    导出访问记录为 CSV 文件

    支持与访问列表相同的筛选条件
    """
    # 构建查询条件（与 get_visits 相同）
    conditions = []

    if device_type:
        conditions.append(Visit.device_type == device_type)

    if min_score is not None:
        conditions.append(Visit.authenticity_score >= min_score)

    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date)
            conditions.append(Visit.timestamp >= start_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="无效的开始日期格式")

    if end_date:
        try:
            end_dt = datetime.fromisoformat(end_date) + timedelta(days=1)
            conditions.append(Visit.timestamp < end_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="无效的结束日期格式")

    # 查询数据
    stmt = select(Visit).order_by(desc(Visit.timestamp)).limit(limit)
    if conditions:
        stmt = stmt.where(and_(*conditions))

    result = await db.execute(stmt)
    visits = result.scalars().all()

    # 创建 CSV
    output = io.StringIO()
    writer = csv.writer(output)

    # 写入表头
    writer.writerow([
        'ID', '访问时间', 'IP地址', '国家', '城市', '设备类型',
        '浏览器', '浏览器版本', '操作系统', 'OS版本',
        '屏幕分辨率', '时区', '语言', '平台',
        '停留时间(秒)', '滚动深度(%)', '鼠标移动',
        '是否机器人', '是否代理', '真实性评分', '指纹哈希',
        '页面URL', '来源'
    ])

    # 写入数据
    for v in visits:
        writer.writerow([
            v.id,
            v.timestamp.isoformat() if v.timestamp else '',
            v.ip_address or '',
            v.ip_country or '',
            v.ip_city or '',
            v.device_type or '',
            v.browser or '',
            v.browser_version or '',
            v.os or '',
            v.os_version or '',
            v.screen_resolution or '',
            v.timezone or '',
            v.language or '',
            v.platform or '',
            v.stay_duration or 0,
            v.scroll_depth or 0,
            v.mouse_movements or 0,
            '是' if v.is_bot else '否',
            '是' if v.is_proxy else '否',
            v.authenticity_score or 0,
            v.fingerprint_hash or '',
            v.page_url or '',
            v.referrer or ''
        ])

    # 返回 CSV 文件
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=visits_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        }
    )


@router.get("/export/json", summary="导出 JSON")
async def export_json(
    device_type: Optional[str] = Query(None, description="设备类型筛选"),
    min_score: Optional[float] = Query(None, ge=0, le=100, description="最低评分"),
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    limit: int = Query(10000, ge=1, le=50000, description="最大导出数量"),
    db: AsyncSession = Depends(get_db)
):
    """
    导出访问记录为 JSON 文件

    支持与访问列表相同的筛选条件
    """
    # 构建查询条件（与 get_visits 相同）
    conditions = []

    if device_type:
        conditions.append(Visit.device_type == device_type)

    if min_score is not None:
        conditions.append(Visit.authenticity_score >= min_score)

    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date)
            conditions.append(Visit.timestamp >= start_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="无效的开始日期格式")

    if end_date:
        try:
            end_dt = datetime.fromisoformat(end_date) + timedelta(days=1)
            conditions.append(Visit.timestamp < end_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="无效的结束日期格式")

    # 查询数据
    stmt = select(Visit).order_by(desc(Visit.timestamp)).limit(limit)
    if conditions:
        stmt = stmt.where(and_(*conditions))

    result = await db.execute(stmt)
    visits = result.scalars().all()

    # 构建 JSON 数据
    data = []
    for v in visits:
        data.append({
            'id': v.id,
            'visit_id': v.visit_id,
            'timestamp': v.timestamp.isoformat() if v.timestamp else None,
            'ip_address': v.ip_address,
            'ip_country': v.ip_country,
            'ip_city': v.ip_city,
            'device_type': v.device_type,
            'browser': v.browser,
            'browser_version': v.browser_version,
            'os': v.os,
            'os_version': v.os_version,
            'screen_resolution': v.screen_resolution,
            'timezone': v.timezone,
            'language': v.language,
            'platform': v.platform,
            'canvas_fingerprint': v.canvas_fingerprint,
            'webgl_fingerprint': v.webgl_fingerprint,
            'fonts_hash': v.fonts_hash,
            'stay_duration': v.stay_duration,
            'scroll_depth': v.scroll_depth,
            'mouse_movements': v.mouse_movements,
            'is_bot': v.is_bot,
            'is_proxy': v.is_proxy,
            'authenticity_score': v.authenticity_score,
            'fingerprint_hash': v.fingerprint_hash,
            'page_url': v.page_url,
            'referrer': v.referrer,
            'user_agent': v.user_agent
        })

    # 返回 JSON 文件
    json_str = json.dumps(data, ensure_ascii=False, indent=2)
    return StreamingResponse(
        iter([json_str]),
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename=visits_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        }
    )


@router.get("/export/excel", summary="导出 Excel")
async def export_excel(
    device_type: Optional[str] = Query(None, description="设备类型筛选"),
    min_score: Optional[float] = Query(None, ge=0, le=100, description="最低评分"),
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    limit: int = Query(10000, ge=1, le=50000, description="最大导出数量"),
    db: AsyncSession = Depends(get_db)
):
    """
    导出访问记录为 Excel 文件

    支持与访问列表相同的筛选条件
    """
    # 构建查询条件（与 get_visits 相同）
    conditions = []

    if device_type:
        conditions.append(Visit.device_type == device_type)

    if min_score is not None:
        conditions.append(Visit.authenticity_score >= min_score)

    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date)
            conditions.append(Visit.timestamp >= start_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="无效的开始日期格式")

    if end_date:
        try:
            end_dt = datetime.fromisoformat(end_date) + timedelta(days=1)
            conditions.append(Visit.timestamp < end_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="无效的结束日期格式")

    # 查询数据
    stmt = select(Visit).order_by(desc(Visit.timestamp)).limit(limit)
    if conditions:
        stmt = stmt.where(and_(*conditions))

    result = await db.execute(stmt)
    visits = result.scalars().all()

    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "访问记录"

    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 定义表头
    headers = [
        'ID', '访问时间', 'IP地址', '国家', '城市', '设备类型',
        '浏览器', '浏览器版本', '操作系统', 'OS版本',
        '屏幕分辨率', '时区', '语言', '平台',
        '停留时间(秒)', '滚动深度(%)', '鼠标移动',
        '是否机器人', '是否代理', '真实性评分', '指纹哈希',
        '页面URL', '来源'
    ]

    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 写入数据
    for row_num, v in enumerate(visits, 2):
        ws.cell(row=row_num, column=1, value=v.id)
        ws.cell(row=row_num, column=2, value=v.timestamp.isoformat() if v.timestamp else '')
        ws.cell(row=row_num, column=3, value=v.ip_address or '')
        ws.cell(row=row_num, column=4, value=v.ip_country or '')
        ws.cell(row=row_num, column=5, value=v.ip_city or '')
        ws.cell(row=row_num, column=6, value=v.device_type or '')
        ws.cell(row=row_num, column=7, value=v.browser or '')
        ws.cell(row=row_num, column=8, value=v.browser_version or '')
        ws.cell(row=row_num, column=9, value=v.os or '')
        ws.cell(row=row_num, column=10, value=v.os_version or '')
        ws.cell(row=row_num, column=11, value=v.screen_resolution or '')
        ws.cell(row=row_num, column=12, value=v.timezone or '')
        ws.cell(row=row_num, column=13, value=v.language or '')
        ws.cell(row=row_num, column=14, value=v.platform or '')
        ws.cell(row=row_num, column=15, value=v.stay_duration or 0)
        ws.cell(row=row_num, column=16, value=v.scroll_depth or 0)
        ws.cell(row=row_num, column=17, value=v.mouse_movements or 0)
        ws.cell(row=row_num, column=18, value='是' if v.is_bot else '否')
        ws.cell(row=row_num, column=19, value='是' if v.is_proxy else '否')
        ws.cell(row=row_num, column=20, value=v.authenticity_score or 0)
        ws.cell(row=row_num, column=21, value=v.fingerprint_hash or '')
        ws.cell(row=row_num, column=22, value=v.page_url or '')
        ws.cell(row=row_num, column=23, value=v.referrer or '')

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # 最大宽度50
        ws.column_dimensions[column_letter].width = adjusted_width

    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 返回 Excel 文件
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=visits_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    )
